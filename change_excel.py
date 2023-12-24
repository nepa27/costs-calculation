from datetime import datetime as dt
from time import ctime
from openpyxl import load_workbook


NAME_BOOK: str = 'costs.xlsx'


def get_data_from_email(data: list) -> list:
    """ Get data from your e-mail. """
    values = []
    for string in data:
        if string.startswith('@'):
            value = string.split('@')
            values.append((int(value[1]), value[2]))
    return values


def get_last_costs() -> list:
    """ Get last costs from costs.xlsx. """
    name_cost = []
    value = []
    book = load_workbook(NAME_BOOK)
    sheet = book.active
    for col in range(1, 3):
        for row in range(2, sheet.max_row + 1):
            if col == 1:
                name_cost.append(sheet.cell(row, col).value)
            if col == 2:
                value.append(sheet.cell(row, col).value)
    last_value = list(zip(name_cost, value))
    return last_value


def write_data(new_values: list, last_values: list) -> None:
    """ Write new data in costs.xlsx. """
    book = load_workbook(NAME_BOOK)
    sheet = book['data']
    for value in new_values:
        cost_value = str(value[0])
        name_cost_value = value[1]
        time = dt.today().strftime("Date: %d.%m.%Y | Time: %H:%M:%S |")
        if value not in last_values:
            sheet.append(value)
            with open('log.txt', 'a') as file:
                file.write(f'| {cost_value} р.'.ljust(11))
                file.write(f'expended for "{name_cost_value}" | ')
                file.write(f'{str(time)}'.rjust(22))
                file.write('\n')

    book.save(NAME_BOOK)
    book.close()


def summ_value_a_week(values_and_numbers: list) -> int:
    """ Get and log costs in a week. """
    summ_coasts = 0
    for value_and_number in values_and_numbers:
        summ_coasts += value_and_number[0]
    return summ_coasts


def delete_data(get_value: list) -> None:
    """ Delete chose items. """
    book = load_workbook(NAME_BOOK)
    sheet = book['data']
    number_rows = [int(value[1:], 16) + 1 for value in get_value]
    print(number_rows)
    [sheet.delete_rows(number) for number in number_rows]
    book.save(NAME_BOOK)
    book.close()


def delete_data_in_the_weekend(last_value: list) -> None:
    """ Delete data about costs in the end of the week. """
    today_date = ctime().split()
    today_day = today_date[0]
    today_time = today_date[3].split(':')
    if today_day == 'Sun' and int(today_time[0]) > 22:
        result_summ = summ_value_a_week(last_value)
        book = load_workbook(NAME_BOOK)
        sheet = book['data']
        sheet.delete_rows(3, 50)
        book.save(NAME_BOOK)
        book.close()
        with open('log.txt', 'a') as file:
            file.write(str(f'Summ for a week: {result_summ}\n {ctime()}\n'))